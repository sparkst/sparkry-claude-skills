#!/usr/bin/env python3
"""
Test Excel Generator Tool (TDD)

REQ-EXCEL-001: Basic file creation from structured data
REQ-EXCEL-002: Conditional formatting (cell colors based on values)
REQ-EXCEL-003: Column width auto-sizing
REQ-EXCEL-004: Freeze panes support
REQ-EXCEL-005: Auto-filter support
REQ-EXCEL-006: Data validation (dropdowns)
REQ-EXCEL-007: Multiple sheets support
"""

import json
import os
import sys
import tempfile
import unittest
from pathlib import Path

# Add the tools directory to path
sys.path.insert(0, str(Path(__file__).parent))

from excel_generator import (
    ExcelGenerator,
    generate_excel,
    apply_conditional_formatting,
    ColorRule,
)


class TestBasicFileCreation(unittest.TestCase):
    """REQ-EXCEL-001: Basic file creation from structured data"""

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.output_path = os.path.join(self.temp_dir, "test_output.xlsx")

    def tearDown(self):
        if os.path.exists(self.output_path):
            os.remove(self.output_path)
        os.rmdir(self.temp_dir)

    def test_creates_xlsx_file_from_list_of_dicts(self):
        """REQ-EXCEL-001 - Creates valid .xlsx file from list of dicts"""
        data = [
            {"Name": "Alice", "Status": "Complete", "Score": 95},
            {"Name": "Bob", "Status": "On Track", "Score": 82},
            {"Name": "Charlie", "Status": "Blocked", "Score": 45},
        ]

        result = generate_excel(data=data, output_path=self.output_path)

        self.assertTrue(os.path.exists(self.output_path))
        self.assertEqual(result["status"], "success")
        self.assertEqual(result["path"], self.output_path)
        self.assertEqual(result["rows"], 3)
        self.assertEqual(result["columns"], 3)

    def test_creates_file_with_headers_from_dict_keys(self):
        """REQ-EXCEL-001 - Headers derived from dict keys"""
        data = [{"Project": "Alpha", "Status": "Active"}]

        generate_excel(data=data, output_path=self.output_path)

        # Verify by reading back
        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Project", headers)
        self.assertIn("Status", headers)
        wb.close()

    def test_returns_error_for_empty_data(self):
        """REQ-EXCEL-001 - Error handling for empty data"""
        result = generate_excel(data=[], output_path=self.output_path)
        self.assertEqual(result["status"], "error")
        self.assertIn("empty", result["message"].lower())

    def test_handles_missing_keys_in_rows(self):
        """REQ-EXCEL-001 - Handles rows with missing keys gracefully"""
        data = [
            {"Name": "Alice", "Status": "Complete"},
            {"Name": "Bob"},  # Missing Status
        ]

        result = generate_excel(data=data, output_path=self.output_path)

        self.assertEqual(result["status"], "success")
        # Should still create file with blank cells for missing data


class TestConditionalFormatting(unittest.TestCase):
    """REQ-EXCEL-002: Conditional formatting (cell colors based on values)"""

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.output_path = os.path.join(self.temp_dir, "test_conditional.xlsx")

    def tearDown(self):
        if os.path.exists(self.output_path):
            os.remove(self.output_path)
        os.rmdir(self.temp_dir)

    def test_applies_color_rules_to_status_column(self):
        """REQ-EXCEL-002 - Status column cells colored based on value"""
        data = [
            {"Task": "Task 1", "Status": "Blocked"},
            {"Task": "Task 2", "Status": "On Track"},
            {"Task": "Task 3", "Status": "Complete"},
        ]

        color_rules = {
            "Status": {
                "Blocked": "FF0000",      # Red
                "On Track": "00FF00",     # Green
                "Complete": "0000FF",     # Blue
            }
        }

        result = generate_excel(
            data=data,
            output_path=self.output_path,
            color_rules=color_rules
        )

        self.assertEqual(result["status"], "success")
        self.assertIn("conditional_formatting_applied", result)
        self.assertTrue(result["conditional_formatting_applied"])

        # Verify colors in file
        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        ws = wb.active

        # Find Status column index
        headers = {cell.value: cell.column for cell in ws[1]}
        status_col = headers["Status"]

        # Check blocked cell (row 2) has red fill
        blocked_cell = ws.cell(row=2, column=status_col)
        self.assertIsNotNone(blocked_cell.fill)
        self.assertEqual(blocked_cell.fill.fgColor.rgb, "FFFF0000")  # ARGB format
        wb.close()

    def test_applies_all_status_colors(self):
        """REQ-EXCEL-002 - All defined status colors applied correctly"""
        data = [
            {"Task": "1", "Status": "Blocked"},
            {"Task": "2", "Status": "Open Issues"},
            {"Task": "3", "Status": "On Track"},
            {"Task": "4", "Status": "Not Started"},
            {"Task": "5", "Status": "Complete"},
            {"Task": "6", "Status": "Canceled"},
        ]

        color_rules = {
            "Status": {
                "Blocked": "FF0000",      # Red
                "Open Issues": "FFFF00",  # Yellow
                "On Track": "00FF00",     # Green
                "Not Started": "808080",  # Grey
                "Complete": "0000FF",     # Blue
                "Canceled": "800080",     # Purple
            }
        }

        result = generate_excel(
            data=data,
            output_path=self.output_path,
            color_rules=color_rules
        )

        self.assertEqual(result["status"], "success")

        # Verify each status has correct color
        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        ws = wb.active
        headers = {cell.value: cell.column for cell in ws[1]}
        status_col = headers["Status"]

        expected_colors = {
            2: "FFFF0000",  # Blocked - Red
            3: "FFFFFF00",  # Open Issues - Yellow
            4: "FF00FF00",  # On Track - Green
            5: "FF808080",  # Not Started - Grey
            6: "FF0000FF",  # Complete - Blue
            7: "FF800080",  # Canceled - Purple
        }

        for row, expected_color in expected_colors.items():
            cell = ws.cell(row=row, column=status_col)
            self.assertEqual(
                cell.fill.fgColor.rgb,
                expected_color,
                f"Row {row} expected {expected_color}"
            )
        wb.close()

    def test_ignores_unknown_values_in_color_rules(self):
        """REQ-EXCEL-002 - Unknown values are not colored"""
        data = [
            {"Task": "1", "Status": "Blocked"},
            {"Task": "2", "Status": "Unknown Status"},
        ]

        color_rules = {
            "Status": {
                "Blocked": "FF0000",
            }
        }

        result = generate_excel(
            data=data,
            output_path=self.output_path,
            color_rules=color_rules
        )

        self.assertEqual(result["status"], "success")
        # Should not error on unknown status values


class TestColumnWidths(unittest.TestCase):
    """REQ-EXCEL-003: Column width auto-sizing"""

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.output_path = os.path.join(self.temp_dir, "test_widths.xlsx")

    def tearDown(self):
        if os.path.exists(self.output_path):
            os.remove(self.output_path)
        os.rmdir(self.temp_dir)

    def test_auto_sizes_columns_to_content(self):
        """REQ-EXCEL-003 - Columns auto-sized to fit content"""
        data = [
            {"Short": "A", "VeryLongColumnHeader": "Value"},
        ]

        result = generate_excel(
            data=data,
            output_path=self.output_path,
            auto_width=True
        )

        self.assertEqual(result["status"], "success")

        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        ws = wb.active

        # Column B (VeryLongColumnHeader) should be wider than Column A
        col_a_width = ws.column_dimensions['A'].width
        col_b_width = ws.column_dimensions['B'].width
        self.assertGreater(col_b_width, col_a_width)
        wb.close()

    def test_custom_column_widths(self):
        """REQ-EXCEL-003 - Custom column widths override auto"""
        data = [{"Name": "Test", "Description": "Value"}]

        column_widths = {"Name": 20, "Description": 50}

        result = generate_excel(
            data=data,
            output_path=self.output_path,
            column_widths=column_widths
        )

        self.assertEqual(result["status"], "success")

        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        ws = wb.active

        headers = {cell.value: cell.column_letter for cell in ws[1]}
        name_col = headers["Name"]
        desc_col = headers["Description"]

        self.assertEqual(ws.column_dimensions[name_col].width, 20)
        self.assertEqual(ws.column_dimensions[desc_col].width, 50)
        wb.close()


class TestFreezePanes(unittest.TestCase):
    """REQ-EXCEL-004: Freeze panes support"""

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.output_path = os.path.join(self.temp_dir, "test_freeze.xlsx")

    def tearDown(self):
        if os.path.exists(self.output_path):
            os.remove(self.output_path)
        os.rmdir(self.temp_dir)

    def test_freezes_header_row_by_default(self):
        """REQ-EXCEL-004 - Header row frozen by default"""
        data = [{"Name": "Alice", "Score": 100}]

        result = generate_excel(
            data=data,
            output_path=self.output_path,
            freeze_panes=True
        )

        self.assertEqual(result["status"], "success")

        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        ws = wb.active
        self.assertEqual(ws.freeze_panes, "A2")  # Freeze below row 1
        wb.close()

    def test_custom_freeze_panes_location(self):
        """REQ-EXCEL-004 - Custom freeze pane location"""
        data = [{"A": 1, "B": 2, "C": 3}]

        result = generate_excel(
            data=data,
            output_path=self.output_path,
            freeze_panes="B3"
        )

        self.assertEqual(result["status"], "success")

        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        ws = wb.active
        self.assertEqual(ws.freeze_panes, "B3")
        wb.close()


class TestAutoFilter(unittest.TestCase):
    """REQ-EXCEL-005: Auto-filter support"""

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.output_path = os.path.join(self.temp_dir, "test_filter.xlsx")

    def tearDown(self):
        if os.path.exists(self.output_path):
            os.remove(self.output_path)
        os.rmdir(self.temp_dir)

    def test_enables_auto_filter_on_all_columns(self):
        """REQ-EXCEL-005 - Auto-filter enabled on all columns"""
        data = [
            {"Name": "Alice", "Status": "Active", "Score": 100},
            {"Name": "Bob", "Status": "Inactive", "Score": 50},
        ]

        result = generate_excel(
            data=data,
            output_path=self.output_path,
            auto_filter=True
        )

        self.assertEqual(result["status"], "success")

        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        ws = wb.active
        self.assertIsNotNone(ws.auto_filter.ref)
        # Should cover all columns (A1:C3 for 3 cols, 3 rows including header)
        self.assertEqual(ws.auto_filter.ref, "A1:C3")
        wb.close()


class TestDataValidation(unittest.TestCase):
    """REQ-EXCEL-006: Data validation (dropdowns)"""

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.output_path = os.path.join(self.temp_dir, "test_validation.xlsx")

    def tearDown(self):
        if os.path.exists(self.output_path):
            os.remove(self.output_path)
        os.rmdir(self.temp_dir)

    def test_adds_dropdown_validation_to_column(self):
        """REQ-EXCEL-006 - Dropdown validation added to Status column"""
        data = [
            {"Task": "Task 1", "Status": "Open"},
            {"Task": "Task 2", "Status": "Closed"},
        ]

        validation_rules = {
            "Status": ["Open", "In Progress", "Closed", "Blocked"]
        }

        result = generate_excel(
            data=data,
            output_path=self.output_path,
            validation_rules=validation_rules
        )

        self.assertEqual(result["status"], "success")

        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        ws = wb.active

        # Check that data validation exists
        self.assertGreater(len(ws.data_validations.dataValidation), 0)
        wb.close()


class TestMultipleSheets(unittest.TestCase):
    """REQ-EXCEL-007: Multiple sheets support"""

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.output_path = os.path.join(self.temp_dir, "test_sheets.xlsx")

    def tearDown(self):
        if os.path.exists(self.output_path):
            os.remove(self.output_path)
        os.rmdir(self.temp_dir)

    def test_creates_multiple_sheets(self):
        """REQ-EXCEL-007 - Creates workbook with multiple sheets"""
        sheets_data = {
            "Summary": [
                {"Project": "Alpha", "Status": "Active"},
                {"Project": "Beta", "Status": "Complete"},
            ],
            "Details": [
                {"Task": "Task 1", "Assignee": "Alice"},
                {"Task": "Task 2", "Assignee": "Bob"},
            ],
        }

        result = generate_excel(
            sheets=sheets_data,
            output_path=self.output_path
        )

        self.assertEqual(result["status"], "success")
        self.assertEqual(result["sheets"], 2)

        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("Details", wb.sheetnames)
        wb.close()

    def test_applies_formatting_per_sheet(self):
        """REQ-EXCEL-007 - Different formatting rules per sheet"""
        sheets_data = {
            "Projects": [
                {"Name": "Alpha", "Status": "Active"},
            ],
            "Tasks": [
                {"Task": "Task 1", "Priority": "High"},
            ],
        }

        sheet_config = {
            "Projects": {
                "color_rules": {
                    "Status": {"Active": "00FF00", "Inactive": "FF0000"}
                }
            },
            "Tasks": {
                "color_rules": {
                    "Priority": {"High": "FF0000", "Low": "00FF00"}
                }
            },
        }

        result = generate_excel(
            sheets=sheets_data,
            output_path=self.output_path,
            sheet_config=sheet_config
        )

        self.assertEqual(result["status"], "success")


class TestExcelGeneratorClass(unittest.TestCase):
    """Test the ExcelGenerator class API"""

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.output_path = os.path.join(self.temp_dir, "test_class.xlsx")

    def tearDown(self):
        if os.path.exists(self.output_path):
            os.remove(self.output_path)
        os.rmdir(self.temp_dir)

    def test_class_builder_pattern(self):
        """Test fluent builder pattern"""
        data = [{"Name": "Test", "Status": "Active"}]

        result = (
            ExcelGenerator(data)
            .with_color_rules({"Status": {"Active": "00FF00"}})
            .with_freeze_panes()
            .with_auto_filter()
            .with_auto_width()
            .save(self.output_path)
        )

        self.assertEqual(result["status"], "success")
        self.assertTrue(os.path.exists(self.output_path))


class TestColorRuleClass(unittest.TestCase):
    """Test ColorRule helper class"""

    def test_creates_color_rule_from_dict(self):
        """ColorRule can be created from dict"""
        rule = ColorRule.from_dict({
            "Status": {
                "Active": "00FF00",
                "Inactive": "FF0000",
            }
        })

        self.assertEqual(rule.get_color("Status", "Active"), "FF00FF00")
        self.assertEqual(rule.get_color("Status", "Inactive"), "FFFF0000")
        self.assertIsNone(rule.get_color("Status", "Unknown"))


if __name__ == "__main__":
    unittest.main()
