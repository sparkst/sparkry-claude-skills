#!/usr/bin/env python3
"""
Excel Generator Tool

Generate Excel files (.xlsx) with advanced formatting including:
- Conditional formatting (cell colors based on values)
- Column width auto-sizing and custom widths
- Freeze panes
- Auto-filter
- Data validation (dropdowns)
- Multiple sheets support

Usage:
    python excel_generator.py --data '<json_data>' --output output.xlsx
    python excel_generator.py --data '<json_data>' --output output.xlsx --color-rules '<json_rules>'

Output (JSON):
    {
        "status": "success",
        "path": "/path/to/output.xlsx",
        "rows": 10,
        "columns": 5,
        "conditional_formatting_applied": true
    }

Dependencies:
    pip install openpyxl
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class ColorRule:
    """Helper class for managing color rules."""

    def __init__(self, rules: Dict[str, Dict[str, str]]):
        """
        Initialize ColorRule with a dict of column->value->color mappings.

        Args:
            rules: Dict like {"Status": {"Active": "00FF00", "Inactive": "FF0000"}}
        """
        self.rules = rules

    @classmethod
    def from_dict(cls, rules: Dict[str, Dict[str, str]]) -> "ColorRule":
        """Create ColorRule from dict."""
        return cls(rules)

    def get_color(self, column: str, value: str) -> Optional[str]:
        """
        Get ARGB color for a column/value pair.

        Args:
            column: Column name
            value: Cell value

        Returns:
            ARGB color string (e.g., "FF00FF00") or None if no rule matches
        """
        if column not in self.rules:
            return None
        color = self.rules[column].get(value)
        if color is None:
            return None
        # Ensure ARGB format (add FF alpha if only RGB provided)
        if len(color) == 6:
            return f"FF{color}"
        return color

    def get_columns(self) -> List[str]:
        """Get list of columns with color rules."""
        return list(self.rules.keys())


class ExcelGenerator:
    """
    Fluent builder for generating Excel files with advanced formatting.

    Example:
        result = (
            ExcelGenerator(data)
            .with_color_rules({"Status": {"Active": "00FF00"}})
            .with_freeze_panes()
            .with_auto_filter()
            .save("output.xlsx")
        )
    """

    def __init__(self, data: Optional[List[Dict[str, Any]]] = None):
        """
        Initialize ExcelGenerator with optional data.

        Args:
            data: List of dicts representing rows
        """
        self.data = data or []
        self.sheets: Dict[str, List[Dict[str, Any]]] = {}
        self.color_rules: Optional[ColorRule] = None
        self.freeze_panes_cell: Optional[str] = None
        self.enable_auto_filter: bool = False
        self.enable_auto_width: bool = False
        self.custom_column_widths: Dict[str, int] = {}
        self.validation_rules: Dict[str, List[str]] = {}
        self.sheet_config: Dict[str, Dict[str, Any]] = {}

    def with_color_rules(self, rules: Dict[str, Dict[str, str]]) -> "ExcelGenerator":
        """
        Add conditional formatting color rules.

        Args:
            rules: Dict like {"Status": {"Active": "00FF00", "Blocked": "FF0000"}}

        Returns:
            self for chaining
        """
        self.color_rules = ColorRule.from_dict(rules)
        return self

    def with_freeze_panes(self, cell: str = "A2") -> "ExcelGenerator":
        """
        Enable freeze panes at specified cell.

        Args:
            cell: Cell reference to freeze at (default "A2" freezes header row)

        Returns:
            self for chaining
        """
        self.freeze_panes_cell = cell
        return self

    def with_auto_filter(self) -> "ExcelGenerator":
        """
        Enable auto-filter on all columns.

        Returns:
            self for chaining
        """
        self.enable_auto_filter = True
        return self

    def with_auto_width(self) -> "ExcelGenerator":
        """
        Enable auto column width sizing.

        Returns:
            self for chaining
        """
        self.enable_auto_width = True
        return self

    def with_column_widths(self, widths: Dict[str, int]) -> "ExcelGenerator":
        """
        Set custom column widths.

        Args:
            widths: Dict like {"Name": 20, "Description": 50}

        Returns:
            self for chaining
        """
        self.custom_column_widths = widths
        return self

    def with_validation(self, rules: Dict[str, List[str]]) -> "ExcelGenerator":
        """
        Add data validation (dropdowns).

        Args:
            rules: Dict like {"Status": ["Open", "Closed", "Blocked"]}

        Returns:
            self for chaining
        """
        self.validation_rules = rules
        return self

    def save(self, output_path: str) -> Dict[str, Any]:
        """
        Generate and save the Excel file.

        Args:
            output_path: Path to save the .xlsx file

        Returns:
            Result dict with status, path, rows, columns
        """
        return generate_excel(
            data=self.data,
            output_path=output_path,
            color_rules=self.color_rules.rules if self.color_rules else None,
            freeze_panes=self.freeze_panes_cell if self.freeze_panes_cell else False,
            auto_filter=self.enable_auto_filter,
            auto_width=self.enable_auto_width,
            column_widths=self.custom_column_widths or None,
            validation_rules=self.validation_rules or None,
        )


def apply_conditional_formatting(
    ws,
    headers: Dict[str, int],
    data_rows: int,
    color_rules: Dict[str, Dict[str, str]],
) -> None:
    """
    Apply cell colors based on value rules.

    Args:
        ws: Worksheet object
        headers: Dict mapping header name to column index
        data_rows: Number of data rows (excluding header)
        color_rules: Dict like {"Status": {"Active": "00FF00"}}
    """
    for column_name, value_colors in color_rules.items():
        if column_name not in headers:
            continue

        col_idx = headers[column_name]

        # Apply color to each cell in the column based on its value
        for row in range(2, data_rows + 2):  # Start from row 2 (after header)
            cell = ws.cell(row=row, column=col_idx)
            cell_value = cell.value

            if cell_value in value_colors:
                color = value_colors[cell_value]
                # Ensure ARGB format
                if len(color) == 6:
                    color = f"FF{color}"
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.fill = fill


def auto_size_columns(ws, headers: Dict[str, str]) -> None:
    """
    Auto-size columns based on content width.

    Args:
        ws: Worksheet object
        headers: Dict mapping header name to column letter
    """
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Add padding and set width
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width


def set_custom_column_widths(
    ws, headers: Dict[str, str], column_widths: Dict[str, int]
) -> None:
    """
    Set custom column widths.

    Args:
        ws: Worksheet object
        headers: Dict mapping header name to column letter
        column_widths: Dict like {"Name": 20, "Description": 50}
    """
    for header_name, width in column_widths.items():
        if header_name in headers:
            col_letter = headers[header_name]
            ws.column_dimensions[col_letter].width = width


def add_data_validation(
    ws,
    headers: Dict[str, int],
    data_rows: int,
    validation_rules: Dict[str, List[str]],
) -> None:
    """
    Add dropdown validation to columns.

    Args:
        ws: Worksheet object
        headers: Dict mapping header name to column index
        data_rows: Number of data rows
        validation_rules: Dict like {"Status": ["Open", "Closed"]}
    """
    for column_name, options in validation_rules.items():
        if column_name not in headers:
            continue

        col_idx = headers[column_name]
        col_letter = get_column_letter(col_idx)

        # Create validation with options list
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"',
            allow_blank=True,
        )
        dv.error = "Please select from the dropdown"
        dv.errorTitle = "Invalid Input"

        # Apply to data cells (not header)
        range_str = f"{col_letter}2:{col_letter}{data_rows + 1}"
        dv.add(range_str)
        ws.add_data_validation(dv)


def generate_excel(
    data: Optional[List[Dict[str, Any]]] = None,
    sheets: Optional[Dict[str, List[Dict[str, Any]]]] = None,
    output_path: str = "output.xlsx",
    color_rules: Optional[Dict[str, Dict[str, str]]] = None,
    freeze_panes: Union[bool, str] = False,
    auto_filter: bool = False,
    auto_width: bool = False,
    column_widths: Optional[Dict[str, int]] = None,
    validation_rules: Optional[Dict[str, List[str]]] = None,
    sheet_config: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """
    Generate an Excel file with advanced formatting.

    Args:
        data: List of dicts for single sheet mode
        sheets: Dict of sheet_name -> list of dicts for multi-sheet mode
        output_path: Path to save the .xlsx file
        color_rules: Conditional formatting rules {"Column": {"Value": "RRGGBB"}}
        freeze_panes: True for "A2", or custom cell reference, or False
        auto_filter: Enable auto-filter on all columns
        auto_width: Auto-size columns to content
        column_widths: Custom column widths {"Column": width}
        validation_rules: Dropdown validation {"Column": ["Option1", "Option2"]}
        sheet_config: Per-sheet config for multi-sheet mode

    Returns:
        Result dict with status, path, rows, columns, etc.
    """
    try:
        # Validate inputs
        if data is None and sheets is None:
            return {
                "status": "error",
                "message": "Either data or sheets must be provided",
            }

        if data is not None and len(data) == 0:
            return {
                "status": "error",
                "message": "Data is empty. Provide at least one row.",
            }

        wb = Workbook()

        # Handle multi-sheet mode
        if sheets is not None:
            # Remove default sheet if we're creating multiple
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            total_rows = 0
            sheet_config = sheet_config or {}

            for sheet_name, sheet_data in sheets.items():
                ws = wb.create_sheet(title=sheet_name)
                config = sheet_config.get(sheet_name, {})

                _populate_sheet(
                    ws=ws,
                    data=sheet_data,
                    color_rules=config.get("color_rules", color_rules),
                    freeze_panes=config.get("freeze_panes", freeze_panes),
                    auto_filter=config.get("auto_filter", auto_filter),
                    auto_width=config.get("auto_width", auto_width),
                    column_widths=config.get("column_widths", column_widths),
                    validation_rules=config.get("validation_rules", validation_rules),
                )
                total_rows += len(sheet_data)

            wb.save(output_path)

            return {
                "status": "success",
                "path": output_path,
                "sheets": len(sheets),
                "total_rows": total_rows,
            }

        # Single sheet mode
        ws = wb.active
        ws.title = "Sheet1"

        rows, cols, has_cf = _populate_sheet(
            ws=ws,
            data=data,
            color_rules=color_rules,
            freeze_panes=freeze_panes,
            auto_filter=auto_filter,
            auto_width=auto_width,
            column_widths=column_widths,
            validation_rules=validation_rules,
        )

        wb.save(output_path)

        result = {
            "status": "success",
            "path": output_path,
            "rows": rows,
            "columns": cols,
        }

        if has_cf:
            result["conditional_formatting_applied"] = True

        return result

    except Exception as e:
        return {
            "status": "error",
            "message": str(e),
        }


def _populate_sheet(
    ws,
    data: List[Dict[str, Any]],
    color_rules: Optional[Dict[str, Dict[str, str]]] = None,
    freeze_panes: Union[bool, str] = False,
    auto_filter: bool = False,
    auto_width: bool = False,
    column_widths: Optional[Dict[str, int]] = None,
    validation_rules: Optional[Dict[str, List[str]]] = None,
) -> tuple:
    """
    Populate a worksheet with data and formatting.

    Returns:
        Tuple of (rows, columns, has_conditional_formatting)
    """
    if not data:
        return 0, 0, False

    # Get all unique keys across all rows for headers
    all_keys = []
    seen_keys = set()
    for row in data:
        for key in row.keys():
            if key not in seen_keys:
                all_keys.append(key)
                seen_keys.add(key)

    # Write headers
    headers_by_name = {}  # name -> column index
    headers_by_letter = {}  # name -> column letter
    for col_idx, header in enumerate(all_keys, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # Style header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        headers_by_name[header] = col_idx
        headers_by_letter[header] = get_column_letter(col_idx)

    # Write data rows
    for row_idx, row_data in enumerate(data, start=2):
        for header, col_idx in headers_by_name.items():
            value = row_data.get(header, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    num_rows = len(data)
    num_cols = len(all_keys)
    has_cf = False

    # Apply conditional formatting
    if color_rules:
        apply_conditional_formatting(ws, headers_by_name, num_rows, color_rules)
        has_cf = True

    # Apply column widths
    if column_widths:
        set_custom_column_widths(ws, headers_by_letter, column_widths)
    elif auto_width:
        auto_size_columns(ws, headers_by_letter)

    # Freeze panes
    if freeze_panes:
        if freeze_panes is True:
            ws.freeze_panes = "A2"
        else:
            ws.freeze_panes = freeze_panes

    # Auto-filter
    if auto_filter:
        last_col = get_column_letter(num_cols)
        last_row = num_rows + 1  # Include header
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Data validation
    if validation_rules:
        add_data_validation(ws, headers_by_name, num_rows, validation_rules)

    return num_rows, num_cols, has_cf


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Generate Excel files with advanced formatting"
    )
    parser.add_argument(
        "--data",
        type=str,
        required=True,
        help="JSON array of objects representing rows",
    )
    parser.add_argument(
        "--output",
        type=str,
        default="output.xlsx",
        help="Output file path (default: output.xlsx)",
    )
    parser.add_argument(
        "--color-rules",
        type=str,
        help='JSON color rules: {"Column": {"Value": "RRGGBB"}}',
    )
    parser.add_argument(
        "--freeze-panes",
        type=str,
        default="A2",
        help="Cell reference for freeze panes (default: A2)",
    )
    parser.add_argument(
        "--auto-filter",
        action="store_true",
        help="Enable auto-filter on all columns",
    )
    parser.add_argument(
        "--auto-width",
        action="store_true",
        help="Auto-size columns to content",
    )

    args = parser.parse_args()

    try:
        data = json.loads(args.data)
    except json.JSONDecodeError as e:
        print(json.dumps({"status": "error", "message": f"Invalid JSON data: {e}"}))
        sys.exit(1)

    color_rules = None
    if args.color_rules:
        try:
            color_rules = json.loads(args.color_rules)
        except json.JSONDecodeError as e:
            print(
                json.dumps(
                    {"status": "error", "message": f"Invalid color rules JSON: {e}"}
                )
            )
            sys.exit(1)

    result = generate_excel(
        data=data,
        output_path=args.output,
        color_rules=color_rules,
        freeze_panes=args.freeze_panes,
        auto_filter=args.auto_filter,
        auto_width=args.auto_width,
    )

    print(json.dumps(result, indent=2))

    if result["status"] == "error":
        sys.exit(1)


if __name__ == "__main__":
    main()
